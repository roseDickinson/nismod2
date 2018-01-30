#!/usr/bin/env python
"""Convert from integration_template.xlsx to yaml project files

Example usage::

    ./excel2yaml.py energy_supply ./integration_template.xlsx ../project/

Arguments
---------
- name of simulation model
- path to template excel file
- path to project directory

Process
-------
- read region_definitions tab
  - write to region_definitions section in project file
- read interval_definitions tab
  - write to interval_definitions section in project file
- read intervals tab
  - write to data folder
  - write to interval csv file in data/interval_definitions
- read units tab
  - write to units in project file
- create new simulation model file under /models
- read sector model inputs tab
  - write to inputs in simulation model file
- read sector model outputs tab
  - write to outputs in simulation model file
- read parameters tab
  - write to parameters in simulation model file
- read interventions tab
  - write to interventions file in data/interventions folder
  - write to interventions section in simulation model file
- read any other tabs
  - write to tab-named yaml file and warn
- TODO write template model_wrapper.py implementing SectorModel
  - get each input and parameter
  - write each output
"""
import csv
import os.path
import sys

# from pprint import pprint

from openpyxl import load_workbook
from ruamel.yaml import YAML


def main(model_name, filename, output_dir):
    """Read then write
    """
    print("Converting data for", model_name)
    print("* Reading existing project data")
    project_data = read_project(output_dir)
    print("* Reading from", filename)
    data = read(model_name, filename, project_data)
    print("* Writing to", output_dir)
    write(model_name, data, output_dir)


def read_project(output_dir):
    """Read existing project data
    """
    try:
        yaml = YAML()
        with open(project_yaml_file(output_dir), encoding='utf-8') as project:
            project_data = yaml.load(project)

        for key, value in project_data.items():
            if value == None:
                project_data[key] = []
        
    except FileNotFoundError:
        project_data = {
            'name': "Test Project",
            'scenario_sets': [],
            'narrative_sets': [],
            'region_definitions': [],
            'interval_definitions': [],
            'units': [],
            'scenarios': [],
            'narratives': []
        }

    return project_data


def project_yaml_file(output_dir):
    """Path to project file
    """
    return os.path.join(output_dir, 'config', 'project.yml')


def read(model_name, filename, project_data):
    """Read workbook to data structure
    """
    workbook = load_workbook(filename=filename)
    expected_worksheets = {
        'region_definitions',
        'interval_definitions',
        'intervals',
        'units',
        'sectormodel_inputs',
        'sectormodel_outputs',
        'parameters',
        'interventions',
        'initial_system',
        'initial_conditions',
        'scenario',
        'scenario_definitions'
    }

    project_data['region_definitions'].extend(read_worksheet('region_definitions', workbook))
    project_data['interval_definitions'].extend(read_worksheet('interval_definitions', workbook))
    project_data['units'].extend(read_worksheet('units', workbook))

    intervals = read_worksheet('intervals', workbook)
    interventions = read_worksheet('interventions', workbook)

    model_data = {
        'name': model_name,
        'description': '{} model description'.format(model_name),
        'path': '/path/to/{}/wrapper.py'.format(model_name),
        'classname': 'SectorModelImplementation',
        'inputs': read_worksheet('sectormodel_inputs', workbook),
        'outputs': read_worksheet('sectormodel_outputs', workbook),
        'parameters': read_worksheet('parameters', workbook),
        'interventions': '{}_interventions.yml'.format(model_name),
        'initial_conditions': read_worksheet('initial_conditions', workbook),
    }

    # Post-process parameter_absolute, relative ranges into tuples
    for parameter in model_data['parameters']:
        parameter['absolute_range'] = '(' + str(parameter['absolute_range_lower']) + ', ' + str(parameter['absolute_range_upper']) + ')'
        parameter['suggested_range'] = '(' + str(parameter['suggested_range_lower']) + ', ' + str(parameter['suggested_range_upper']) + ')'

        del parameter['absolute_range_lower']
        del parameter['absolute_range_upper']
        del parameter['suggested_range_lower']
        del parameter['suggested_range_upper']

    extra = {}
    for sheet_name in set(workbook.sheetnames) - expected_worksheets:
        print("  * Found extra worksheet", sheet_name)
        extra[sheet_name] = read_worksheet(sheet_name, workbook)

    return project_data, intervals, interventions, model_data, extra


def read_worksheet(worksheet_name, workbook):
    """Read worksheet table to list of dicts
    """
    output = []
    try:
        data = workbook[worksheet_name].values
    except KeyError:
        print("Worksheet {} not found".format(worksheet_name))
        return output

    keys = next(data)
    data = list(data)

    for row in data:
        item = dict()
        for key, value in zip(keys, row):
            if key is not None:
                item[key] = value
            else:
                if value is not None:
                    print(worksheet_name, value, "had no column name")
        output.append(item)

    return output


def write(model_name, data, output_dir):
    """Write data structure to YAML and csv
    """
    project_data, intervals, interventions, model_data, extra = data

    yaml = YAML()

    # project
    with open(project_yaml_file(output_dir), 'w', encoding='utf-8') as project_file:
        yaml.dump(project_data, project_file)

    # intervals
    intervals_filename = os.path.join(output_dir, 'data', 'interval_definitions','{}_intervals.csv'.format(model_name))
    with open(intervals_filename, 'w', encoding='utf-8', newline='') as intervals_file:
        fieldnames = ('id', 'start_hour', 'end_hour')
        writer = csv.DictWriter(intervals_file, fieldnames)
        writer.writeheader()
        writer.writerows(intervals)

    # interventions
    interventions_filename = os.path.join(output_dir, 'data', 'interventions', '{}_interventions.yml'.format(model_name))
    with open(interventions_filename, 'w', encoding='utf-8') as interventions_file:
        yaml.dump(interventions, interventions_file)

    # model
    model_filename = os.path.join(
        output_dir, 'config', 'sector_models', '{}.yml'.format(
            model_name))
    with open(model_filename, 'w', encoding='utf-8') as model_file:
        yaml.dump(model_data, model_file)

    

    # extras
    for sheet_name, data in extra.items():
        filename = os.path.join(output_dir, '{}__{}.yml'.format(model_name, sheet_name))
        with open(filename, 'w', encoding='utf-8') as file_handle:
            yaml.dump(data, file_handle)

if __name__ == '__main__':
    if len(sys.argv) != 4:
        SCRIPT_NAME = os.path.basename(__file__)
        print(
            "Usage: python {} model_name integration_template.xslx path/to/project/dir/".format(
                SCRIPT_NAME))
        exit(1)

    MODEL_NAME = sys.argv[1]
    FILENAME = os.path.normpath(os.path.abspath(sys.argv[2]))
    OUTPUT_DIR = os.path.normpath(os.path.abspath(sys.argv[3]))
    main(MODEL_NAME, FILENAME, OUTPUT_DIR)
